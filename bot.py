import PySimpleGUI as sg
import subprocess
import requests
from openpyxl import load_workbook, Workbook

# Função para pesquisar e capturar os dados do CNPJ
def pesquisar_cnpj(cnpj):
    url = f"https://www.receitaws.com.br/v1/cnpj/{cnpj}"
    response = requests.get(url)
    data = response.json()

    # Capturar os dados relevantes
    nome_empresa = data["nome"]
    nome_fantasia = data["fantasia"]
    endereco = f"{data['logradouro']}, {data['numero']} - {data['bairro']}"
    telefone = data["telefone"]
    email = data["email"]
    area_atuacao = data["atividade_principal"][0]["text"]

    # Retornar os dados capturados
    return [nome_empresa, nome_fantasia, endereco, telefone, email, area_atuacao]

if __name__ == "__main__":
    try:
        # Carregar a planilha existente
        planilha = load_workbook("dados_empresas.xlsx")
        print("Planilha existente carregada com sucesso!")
    except FileNotFoundError:
        # Caso o arquivo não exista, criar uma nova planilha vazia
        planilha = Workbook()
        planilha.active.append(["Nome da Empresa", "Nome Fantasia", "Endereço", "Telefone", "Email", "Área de Atuação"])
        print("Nova planilha criada!")

    sheet = planilha.active

    # Verificar se o código CNPJ foi passado como argumento
    import sys
    if len(sys.argv) > 1:
        cnpj = sys.argv[1]
        dados_cnpj = pesquisar_cnpj(cnpj)
        sheet.append(dados_cnpj)

        # Ajustar o tamanho das células
        for column_cells in sheet.columns:
            max_length = 0
            for cell in column_cells:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column_cells[0].column_letter].width = adjusted_width

        # Salvar a planilha em um arquivo
        planilha.save("dados_empresas.xlsx")
        print("Dados salvos na planilha com sucesso!")

layout = [
    [sg.Text("Digite o CNPJ para pesquisa (apenas números).")],
    [sg.InputText(key="cnpj")],
    [sg.Button("Pesquisar"), sg.Button("Sair")],
    [sg.Text("", key="response")]
]

window = sg.Window("Search Enterprise", layout)

while True:
    event, values = window.read()
    if event == sg.WINDOW_CLOSED or event == "Sair":
        break
    if event == "Pesquisar":
        cnpj = values["cnpj"]
        response = subprocess.run(["python", "bot.py", cnpj], capture_output=True, text=True)
        output = response.stdout.strip()
        window["response"].update(output)

window.close()